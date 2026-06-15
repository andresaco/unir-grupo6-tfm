import os

import pandas as pd
import yfinance as yf
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill


def ejecutar_desarrollo():
    """
    Ejecuta el flujo completo de adquisición y reporte de datos bursátiles.

    Descarga datos históricos de Apple (AAPL) del último mes, calcula indicadores
    técnicos básicos (Evolución % y Media Móvil), extrae datos fundamentales
    y genera un archivo Excel con formato y gráficos integrados.

    Raises:
        Exception: Si ocurre un error durante la descarga o el procesamiento de datos.
    """
    try:
        ticker_symbol = "AAPL"
        print(f"--- Iniciando proceso para {ticker_symbol} ---")

        # 1. DESCARGA DE DATOS (Último mes)
        apple = yf.Ticker(ticker_symbol)
        df = apple.history(period="1mo")

        if df.empty:
            print("Error: No se pudieron descargar datos de la fuente.")
            return

        # 2. PROCESAMIENTO TÉCNICO
        df.index = df.index.strftime("%Y-%m-%d")
        df = df.reset_index()
        # Evolución diaria
        df["Evolucion_%"] = ((df["Close"] - df["Open"]) / df["Open"] * 100).round(2)
        # Media Móvil (SMA 10)
        df["Media_Movil_10"] = df["Close"].rolling(window=10).mean().round(2)

        # 3. DATOS FUNDAMENTALES
        info = apple.info
        fundamentales = {
            "Métrica": ["Capitalización", "PER", "Precio Obj.", "Dividendo", "Margen"],
            "Valor": [
                f"{info.get('marketCap', 0):,}",
                info.get("trailingPE", "N/A"),
                info.get("targetMeanPrice", "N/A"),
                f"{info.get('dividendYield', 0) * 100:.2f}%"
                if info.get("dividendYield")
                else "0%",
                f"{info.get('profitMargins', 0) * 100:.2f}%",
            ],
        }
        df_fun = pd.DataFrame(fundamentales)

        # 4. CREACIÓN DEL EXCEL (.xlsx)
        nombre_archivo = "Reporte_Final_Apple.xlsx"

        with pd.ExcelWriter(nombre_archivo, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Datos_Tecnicos")
            df_fun.to_excel(writer, index=False, sheet_name="Fundamentales")

            # Diseño y Gráficos
            ws = writer.sheets["Datos_Tecnicos"]

            # Estilo de cabecera
            for cell in ws[1]:
                cell.fill = PatternFill(
                    start_color="131722", end_color="131722", fill_type="solid"
                )
                cell.font = Font(color="FFFFFF", bold=True)

            # Insertar Gráfico de Precio y Volumen
            chart_precio = LineChart()
            chart_precio.title = "AAPL: Precio vs Volumen"
            chart_precio.style = 13

            # Datos de Cierre (Columna 5: Close)
            precios = Reference(ws, min_col=5, min_row=1, max_row=len(df) + 1)
            fechas = Reference(ws, min_col=1, min_row=2, max_row=len(df) + 1)

            chart_precio.add_data(precios, titles_from_data=True)
            chart_precio.set_categories(fechas)

            # Añadir Volumen (Columna 6: Volume) en eje secundario
            chart_vol = BarChart()
            volumen = Reference(ws, min_col=6, min_row=1, max_row=len(df) + 1)
            chart_vol.add_data(volumen, titles_from_data=True)
            chart_vol.y_axis.axId = 200
            chart_vol.y_axis.crosses = "max"

            chart_precio += chart_vol
            ws.add_chart(chart_precio, "J2")

        print(f"¡ÉXITO! Documento creado correctamente en: {os.getcwd()}")
        print(f"Nombre del archivo: {nombre_archivo}")

    except Exception as e:
        print(f"Hubo un error al ejecutar: {e}")


if __name__ == "__main__":
    ejecutar_desarrollo()
